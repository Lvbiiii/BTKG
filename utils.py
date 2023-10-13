'''
各种文件读写的公共函数
'''
import json
import pandas as pd
import numpy as np
import csv
import openpyxl
from openpyxl import load_workbook

def check_charset(file_path):
    import chardet
    with open(file_path, "rb") as f:
        data = f.read(4)
        charset = chardet.detect(data)['encoding']
    return charset
def readJSON(fileName):
    f=open(fileName,'r',encoding=check_charset(fileName),errors='ignore')
    content=f.read()
    data=json.loads(content,strict=False)
    return data

def readCSV(filename):
    df=pd.read_csv(filename)
    return df

def writeCSV(filename,data,header):
    with open(filename,'a',newline='',encoding='utf-8') as file:
        writer=csv.writer(file)
        if header is not None:
            writer.writerow(header)
        for d in data:
            writer.writerow(d)
def readExcel(filename,myrow,mycolumn):
    table = openpyxl.load_workbook(filename)['Sheet']
    data = table.cell(row=myrow,column=mycolumn).value
    return data

def writeExcel(filename,myrow,mycolumn,data):
    wb = load_workbook(filename)
    sheet = wb.active
    sheet.cell(row=myrow,column=mycolumn).value = data
    wb.save(filename)

def saveNpy(filename,data):
    np.save(filename,data)

def loadNpy(filename):
    return np.load(filename,allow_pickle=True)

def readTxt(filename):
    f=open(filename)
    return f.readlines()
